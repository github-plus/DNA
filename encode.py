import xlrd
import xlwt
from openpyxl import Workbook
import config
from todna import todna


# 看懂代码和思路后才能改以下三个参数，随意更改将出现致命级错误！！！
# 看懂代码和思路后才能改以下三个参数，随意更改将出现致命级错误！！！
# 看懂代码和思路后才能改以下三个参数，随意更改将出现致命级错误！！！
#假设位置信息的nt数位7，字符对应的nt数为8，每条连存储的字符个数为6，那么每条链的长度为7+8*6=55
char_nt = config.char_nt  #在此设置字库中每个字符对应的nt数
address_nt = config.address_nt #在此设置位置信息的nt数
char_num = config.char_num #在此设置每条连链中存储的字符个数
range_num = config.range_num  #迭代次数
code_note = config.code_note #每编码n条打印一次
max_replace = config.max_replace #最大被替换的字符个数
replace_at = config.replace_at #替换字符的模板，比如说1个字符7个nt
replace_gc = config.replace_gc #替换字符的模板，比如说1个字符7个nt
font = config.font     # 编码使用的字库
char_bit = config.char_bit
# 看懂代码和思路后才能改以上三个参数，随意更改将出现致命级错误！！！
# 看懂代码和思路后才能改以上三个参数，随意更改将出现致命级错误！！！
# 看懂代码和思路后才能改以上三个参数，随意更改将出现致命级错误！！！
file_log = config.log_encode
nums = todna.ntNum(address_nt)
global err1,err2
err1 = []
err2 = [] #改造后还出现gc<0.4的链
sum = 0 #被替换的字符个数

#使用xlsx格式，最大行为100w
workbook = Workbook()
sheet = workbook.active
# 日志文件，用以存储找不到的字符
log = open(file_log,mode='w',encoding='utf-8')
'''
# 创建一个workbook 设置编码
workbook = xlwt.Workbook(encoding='utf-8')
sheet = workbook.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
'''
def encode(filename):


    # 需要读取的文档
    file = f"doc/{filename}.txt"


    # 读取字符库
    workBook = xlrd.open_workbook(font)
    char_to_dna = workBook.sheet_by_index(0)
    global result,char_address
    result = {}

    # 获取总共行数
    rows = char_to_dna.nrows
    print("正在开始将字库写入字典中")
    for i in range(rows):
        cha = char_to_dna.cell_value(i,0)
        dna = char_to_dna.cell_value(i,1)
        result[cha]=dna
    print("字典写入完成")


    #读取文档内容
    print("开始读取文档内容")
    try:
        with open(file, "r",encoding='utf-8') as f:  # 打开文件
            data = f.read()  # 读取文件
    except:
        print("该文件不存在，请重新输入文件")
        return False
    print("文档内容读取完毕")

    #开始对文档内容进行DNA编码
    num = 0 #此为第n个位置信息
    temp = ''


    #开始对文档内容进行编码为DNA

    for content in data:
        temp = temp+content
        #累计为设置好的1每条链字符个数才开始进行编码
        if len(temp) == char_num:
            #每进行100个编码提示一次
            if num %code_note == 0:
                print("正在进行第",num,"条链编码")

            temp = transform(num,temp)
            num += 1
            #如果num的个数大于pow(4,address_nt),将要停止进行编码
            if num > pow(4,address_nt):
                print(f"{address_nt}nt数量不够存储本文件，请增加nt数")

    #如果最后字符个数不足以凑齐，将使用空格代替
    if temp != '':
        for i in range(char_num-len(temp)):
            temp +=' '
        temp = transform(num, temp)
    data_len = len(data)  # 输入数据的长度
    #char_bit = 8  # 以gb2313编码为参考，1个字符占16bits
    strand_len = char_num*char_nt+address_nt

    print(f"文档转化为DNA链完成,总共生成了{num}条链")
    print(f"总共替换了{sum}个字符，损耗率为{sum/data_len}")
    print(f"编码潜力为{(data_len*char_bit)/(num*strand_len)}bit/nt")
    print(f"总的二进制个数为{data_len*char_bit}，字节数为{data_len*char_bit/8}")
    print("对于没有找到的字符请查看char_encode.log")
    workbook.save(f'doc/{filename}.xlsx')
    return True

def transform(num,temp):
    back_char = '' #需要返回的字符
    #获取到位置信息的DNA序列
    address_strand = nums[num]
   # print(f"{num}对应的DNA链为{address_strand}")
    content_strand = '' #字符转换后的DNA链

    # print(temp)
    for i in temp:
        # print(f"{i}对应的DNA序列为{result.get(i)}")
        try:
            content_strand += result.get(i)
        except:
            print(f"{i}在字库中找不到对应的DNA序列")
            log.write(i)
    strand = address_strand+content_strand
    # print(strand)
    #print(f"整链为{strand}")
    last = encrypt(num,strand) #未进行转化的字符
    end = len(temp)-int(last) #最后一个被转换的字符
    if int(last) !=0:
        last = int('-' + str(last))
        back_char = temp[last:]
        # print(f"需要返回的字符串是{back_char}")
    '''
    #将字符写入excle
    try:
        sheet.write(num,0,temp[:end])
    except:
        print("{temp[:end]}字符插入失败")
    '''
    return back_char


def encrypt(num,strand):
    new_strand = strand
    global sum
    last = 0
    #迭代次数
    for i in range(0,range_num):
        for j in range(len(new_strand)):
            index = (i+2)*j
            if index < len(new_strand):

                if new_strand[index] == 'A':
                    new_strand = new_strand[:index] + 'C' + new_strand[index+1:]
                elif new_strand[index] == 'C':
                    new_strand = new_strand[:index] + 'T' + new_strand[index+1:]
                elif new_strand[index] == 'T':
                    new_strand = new_strand[:index] + 'G' + new_strand[index+1:]
                elif new_strand[index] == 'G':
                    new_strand = new_strand[:index] + 'A' + new_strand[index+1:]

    percentage_1 = gc(new_strand) #改造前的gc含量
    if percentage_1 < 0.4 or percentage_1 > 0.6 :
        temp_new_strand = ''
        if percentage_1 > 0.6:

            # 替换字符所对应的nt数，假如1个字符对应的char_num=7，替换1个字符，那就是替换最末尾7个nt；替换2个字符，那就是替换最末未14个nt
            for replace_num in range(1,max_replace+1):
                replace_nt_at = ''
                for i in range(replace_num):
                    replace_nt_at += replace_at
                sum += 1
                last = replace_num
                temp_new_strand = new_strand[:adverse(char_nt*replace_num)] + replace_nt_at
                if gc(temp_new_strand) < 0.6:
                    break

            '''
            # 替换1个字符所对应的nt数，假如1个字符对应的char_num=7，那就是替换最末尾7个nt
            temp_new_strand = new_strand[:adverse(char_num)] + 'ATATATA'
            sum +=1
            last = 1
            if gc(temp_new_strand) > 0.6:
                # 替换2个字符所对应的nt数，假如1个字符对应的char_num=7，那就是替换最末尾14个nt
                temp_new_strand = new_strand[:adverse(char_num*2)] + 'ATATATAATATATA'
                sum += 1
                last=2
                if gc(temp_new_strand) > 0.6:
                    temp_new_strand = new_strand[:adverse(char_num*3)] + 'ATATATAATATATAATATATA'
                    sum += 1
                    last = 3
            '''
        if percentage_1 < 0.4:


            for replace_num in range(1, max_replace + 1):
                replace_nt_gc = ''
                for i in range(replace_num):
                    replace_nt_gc += replace_gc
                sum += 1
                last = replace_num
                temp_new_strand = new_strand[:adverse(char_nt*replace_num)] + replace_nt_gc
                if gc(temp_new_strand) > 0.4:
                    break
            '''
            temp_new_strand = new_strand[:adverse(char_num)] + 'GCGCGCG'
            sum += 1
            last = 1
            if gc(temp_new_strand)  < 0.4:
                temp_new_strand = new_strand[:adverse(char_num*2)] + 'GCGCGCGGCGCGCG'
                sum += 1
                last = 2
                if gc(temp_new_strand)  < 0.4:
                    temp_new_strand = new_strand[:adverse(char_num*3)] + 'GCGCGCGGCGCGCGGCGCGCG'
                    sum += 1
                    last = 3
            '''
        percentage_2 = gc(temp_new_strand) #改造后的gc含量
        if percentage_2 < 0.4 or percentage_2 > 0.6:

            log.write(f"改造失败，具体链为{new_strand}，转换后为{temp_new_strand},转换后的gc含量为{percentage_2}")
            print("改造前", new_strand)
            print("改造前AT含量为", percentage_1)
            print("改造后AT含量为", percentage_2)
            print("改造后", temp_new_strand)
            err2.append(new_strand)
        new_strand = temp_new_strand

    try:
        # sheet.write(num,1,new_strand)
        #sheet.write(num, 0, new_strand) #此为xlwt用法

        sheet.cell(num+1,1).value = new_strand #此为xlsx用法
    except:
        print(f"{new_strand}插入excle文档失败")
    return last

# 计算gc含量
def gc(strand):

    gc =  (strand.count('G') + strand.count('C')) / (strand.count('A') + strand.count('T') + strand.count('C') + strand.count('G'))
    return gc

# 正整数取反取反函数
def adverse(digit):
    reverse_digit = '-' + str(digit)
    return int(reverse_digit)