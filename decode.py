from todna import todna
from openpyxl import load_workbook
import xlrd
import  match
import BaseUtils
import config
# 看懂代码和思路后才能改以下三个参数，随意更改将出现致命级错误！！！
# 看懂代码和思路后才能改以下三个参数，随意更改将出现致命级错误！！！
# 看懂代码和思路后才能改以下三个参数，随意更改将出现致命级错误！！！
#假设位置信息的nt数位7，字符对应的nt数为8，每条连存储的字符个数为6，那么每条链的长度为7+8*6=55
char_nt = config.char_nt  #在此设置字库中每个字符对应的nt数
address_nt = config.address_nt #在此设置位置信息的nt数
char_num = config.char_num #在此设置每条连链中存储的字符个数
range_num = config.range_num-1  #迭代次数,这个数为encode中range_num-1
decode_note = config.code_note #每编码n条打印一次
max_replace = config.max_replace   #最大被替换的字符数，一般以总字符数在2/5最好。比如说设计的1条为链51个nt，那么51*0.4=20.4。7个nt对应1个字符，其中最大被替换字符最好是3个，因为此时最大被替换nt数是7*3=21nt，接近20.4nt。
font = config.font     # 编码使用的字库
replace_at = config.replace_at
replace_gc = config.replace_gc
file_log = config.log_decode

# 看懂代码和思路后才能改以上三个参数，随意更改将出现致命级错误！！！
# 看懂代码和思路后才能改以上三个参数，随意更改将出现致命级错误！！！
# 看懂代码和思路后才能改以上三个参数，随意更改将出现致命级错误！！！


log = open(file_log,mode='w',encoding='utf-8')
nums = todna.ntNum(address_nt)

dic = {}
def decode(filename):

    # 读取字符库
    workBook = xlrd.open_workbook(font)
    # 获取读取的数据
    char_to_dna = workBook.sheet_by_index(0)
    global result
    result = {}

    # 获取总共行数
    rows = char_to_dna.nrows
    print("正在开始将字库写入字典中")
    for i in range(rows):
        cha = char_to_dna.cell_value(i, 0)
        dna = char_to_dna.cell_value(i, 1)
        result[cha] = dna
    print("字典写入完成")

    #此版本为后缀xls
    ''' 
    # 读取excle中strand信息
    workbook = xlrd.open_workbook(f'doc/{filename}.xls')
    sheet = workbook.sheet_by_index(0)
    '''
    # 此版本为后缀xlsx
    workbook = load_workbook(f'doc/{filename}.xlsx')
    sheet = workbook.active

    print("正在开始解码")
    # 遍历每一行，获得每条strand
    for row in range(1,sheet.max_row+1):
        strand = sheet.cell(row,1).value
        address = strand[:address_nt]
        char = strand[address_nt:]
        decrypt(strand)
        if row%decode_note == 0:
            print(f"正在解码到第{row}链,总共有{sheet.max_row}条")
    '''
    print("正在开始解码")
    #遍历每一行，获得每条strand
    for row in range(sheet.nrows):
        strand = sheet.cell_value(row,0)
        address = strand[:address_nt]
        char = strand[address_nt:]
        decrypt(strand)
    '''
    words = '' #解码后的文档内容
    # 对获取到的数据进行排序，以获得正确的字符排序
    dic_num = sorted(dic.keys())
    for key in dic_num:
        words += dic[key]
    print("解码完成，数据为:")
    print(words)
    filename = filename.replace('encode','decode')
    with open(f'doc/{filename}.txt',mode='w',encoding='utf-8') as f:
        f.write(words)
        f.close()



def decrypt(strand):

    new_strand = strand
    #  警告：解码迭代参数要和编码迭代参数一样
    #  警告：解码迭代参数要和编码迭代参数一样
    #  警告：解码迭代参数要和编码迭代参数一样
    last = lastchar(new_strand)
    for i in range(range_num, -1, -1):

        for j in range(len(new_strand)):
            index = (i + 2) * j
            if index < len(new_strand):
                if new_strand[index] == 'C':
                    new_strand = new_strand[:index] + 'A' + new_strand[index + 1:]
                elif new_strand[index] == 'T':
                    new_strand = new_strand[:index] + 'C' + new_strand[index + 1:]
                elif new_strand[index] == 'G':
                    new_strand = new_strand[:index] + 'T' + new_strand[index + 1:]
                elif new_strand[index] == 'A':
                    new_strand = new_strand[:index] + 'G' + new_strand[index + 1:]
    # 取得位置信息的DNA序列，并将其转化为十进制
    address = new_strand[:address_nt]
    # print(f"取得位置信息的DNA序列{address}")
    address = match.dna_to_num(address)
    address = BaseUtils.four_to_ten(address)
    # print(f"取得位置信息的十进制{address}")

    #取得正确数据的DNA序列
    end = len(new_strand)-last*char_nt
    data = new_strand[address_nt:end]
    # print(f"取得数据信息的DNA序列{data}")
    #根据码表，每个DNA序列解码为字符

    word = ''
    for j in range(char_num-last):
        word_dna = data[j*char_nt:(j+1)*char_nt]
        try:

            word += list(result.keys())[list(result.values()).index(word_dna)]
        except:
            word +="*"
            print(data)
            print(f"{word_dna}不能在码表中找到")
            log.write(word_dna)
    # print(word)
    dic[address] = word


def lastchar(strand):
    # 警告：截取参数需要和编码替换参数一致，可以在encode类的encrypt函数中查看
    # 警告：截取参数需要和编码替换参数一致，可以在encode类的encrypt函数中查看
    # 警告：截取参数需要和编码替换参数一致，可以在encode类的encrypt函数中查看
    # 查看该链被替换了nt数
    for replace_num in range(max_replace,-1,-1):
        replace_nt_at = ''
        replace_nt_gc = ''
        if replace_num == 0:
            return 0
        for i in range(replace_num):
            replace_nt_at += replace_at
            replace_nt_gc += replace_gc
        if strand[adverse(replace_num*char_nt):] == replace_nt_at or strand[adverse(replace_num*char_nt):] == replace_nt_gc:
            return replace_num
    '''
    原版
    if strand[-21:] =='ATATATATATATATATATATA' or strand[-21:] == 'CGCGCGCGCGCGCGCGCGCGC':
        return 3
    elif strand[-14:] == 'ATATATATATATAT' or strand[-14:] == 'CGCGCGCGCGCGCG':
        return 2
    elif strand[-7:] == 'TATATAT' or strand[-7:] == 'GCGCGCG':
        return 1
    else:
        return 0
    '''
# 正整数取反取反函数
def adverse(char_num):
    char_num_num = '-' + str(char_num)
    return int(char_num_num)